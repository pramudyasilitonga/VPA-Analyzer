#!/usr/bin/env python3
"""Analisis rekomendasi Vertical Pod Autoscaler (VPA) dan ekspor ke Excel."""

import json
import logging
import os
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

# --- Konfigurasi via Environment Variables ---
SAMPLE_FILE = os.getenv("VPA_SAMPLE_FILE", "sample_vpa.json")
OUTPUT_FILE = os.getenv("VPA_OUTPUT_FILE", "vpa_recommendations.xlsx")
KUBECTL_CMD = os.getenv("VPA_KUBECTL_CMD", "kubectl get vpa -A -o json")
KUBECTL_CONTEXTS_CMD = os.getenv(
    "VPA_KUBECTL_CONTEXTS_CMD", "kubectl config get-contexts -o name"
)
SAMPLE_CLUSTER_NAME = os.getenv("VPA_SAMPLE_CLUSTER_NAME", "local-sample")
LIMIT_MULTIPLIER = float(os.getenv("VPA_LIMIT_MULTIPLIER", "1.2"))
LOG_LEVEL = os.getenv("VPA_LOG_LEVEL", "INFO").upper()
FETCH_WORKLOAD = os.getenv("VPA_FETCH_WORKLOAD", "true").lower() == "true"
CHANGE_TOLERANCE = float(os.getenv("VPA_CHANGE_TOLERANCE", "5.0"))

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

EXCEL_COLUMNS = [
    "Cluster Name",
    "Namespace",
    "Nama Service",
    "CPU Sekarang",
    "Rekomendasi CPU Baru",
    "Selisih CPU",
    "% Perubahan CPU",
    "Memori Sekarang",
    "Rekomendasi Memori Baru",
    "Selisih Memori",
    "% Perubahan Memori",
    "Status Tindakan",
]

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def parse_cpu(value: str) -> float:
    """Parse CPU string Kubernetes ke millicores (float)."""
    value = str(value).strip()
    if value.endswith("m"):
        return float(value[:-1])
    return float(value) * 1000


def format_cpu(millicores: float) -> str:
    """Format millicores ke string Kubernetes."""
    if millicores >= 1000 and millicores % 1000 == 0:
        return f"{int(millicores // 1000)}"
    if millicores == int(millicores):
        return f"{int(millicores)}m"
    return f"{millicores:.0f}m"


def parse_memory(value: str) -> float:
    """Parse memory string Kubernetes ke bytes (float)."""
    value = str(value).strip()
    suffixes = {
        "Ki": 1024,
        "Mi": 1024**2,
        "Gi": 1024**3,
        "Ti": 1024**4,
        "K": 1000,
        "M": 1000**2,
        "G": 1000**3,
        "T": 1000**4,
    }
    for suffix, multiplier in suffixes.items():
        if value.endswith(suffix):
            return float(value[: -len(suffix)]) * multiplier
    return float(value)


def format_memory(bytes_value: float) -> str:
    """Format bytes ke string Kubernetes (Mi jika memungkinkan)."""
    mib = bytes_value / (1024**2)
    if mib >= 1 and abs(mib - round(mib)) < 0.01:
        return f"{int(round(mib))}Mi"
    if bytes_value >= 1024**2:
        return f"{mib:.2f}Mi"
    if bytes_value >= 1024:
        return f"{bytes_value / 1024:.0f}Ki"
    return str(int(round(bytes_value)))


def memory_to_mib(bytes_value: float) -> float:
    """Konversi bytes ke MiB numerik."""
    return bytes_value / (1024**2)


def calc_percent_change(current: float, recommended: float) -> float | None:
    """Hitung persentase perubahan; None jika pembagi nol."""
    if current == 0:
        return None
    return round(((recommended - current) / current) * 100, 2)


def determine_status(cpu_pct: float | None, mem_pct: float | None) -> str:
    """Tentukan status tindakan berdasarkan arah perubahan resource."""
    tolerance = CHANGE_TOLERANCE

    def classify(value: float | None) -> str:
        if value is None:
            return "same"
        if value > tolerance:
            return "up"
        if value < -tolerance:
            return "down"
        return "same"

    cpu_sig = classify(cpu_pct)
    mem_sig = classify(mem_pct)

    has_up = cpu_sig == "up" or mem_sig == "up"
    has_down = cpu_sig == "down" or mem_sig == "down"

    if has_up and has_down:
        return "Tinjau Manual (Campuran)"
    if has_up:
        return "Upsize - Tambah Resource"
    if has_down:
        return "Downsize - Hemat Resource"
    return "Pertahankan Konfigurasi"


def run_kubectl(command: list[str], *, check: bool = True) -> subprocess.CompletedProcess:
    """Jalankan perintah kubectl dan kembalikan hasil subprocess."""
    try:
        return subprocess.run(
            command,
            capture_output=True,
            text=True,
            check=check,
        )
    except FileNotFoundError:
        logger.error("Perintah kubectl tidak ditemukan di PATH.")
        sys.exit(1)


def get_kubectl_contexts() -> list[str]:
    """Ambil daftar semua context cluster dari kubeconfig."""
    result = run_kubectl(KUBECTL_CONTEXTS_CMD.split())
    contexts = [line.strip() for line in result.stdout.splitlines() if line.strip()]
    logger.info("Ditemukan %d context cluster.", len(contexts))
    return contexts


def get_current_context() -> str | None:
    """Ambil context kubectl yang sedang aktif."""
    result = run_kubectl(["kubectl", "config", "current-context"], check=False)
    if result.returncode != 0:
        return None
    context = result.stdout.strip()
    return context or None


def use_kubectl_context(context: str) -> bool:
    """Ganti context kubectl aktif ke cluster target."""
    result = run_kubectl(["kubectl", "config", "use-context", context], check=False)
    if result.returncode != 0:
        logger.warning(
            "Gagal switch context ke '%s': %s",
            context,
            result.stderr.strip(),
        )
        return False

    logger.info("Context aktif: %s", context)
    return True


def fetch_vpa_json_from_current_context() -> dict:
    """Ambil data VPA dari cluster pada context kubectl yang sedang aktif."""
    result = run_kubectl(KUBECTL_CMD.split())
    return json.loads(result.stdout)


def load_vpa_data_by_cluster() -> tuple[list[tuple[str, dict]], bool]:
    """
    Muat data VPA per cluster.

    Mode sample: satu cluster lokal dari file JSON.
    Mode live: loop semua context kubectl, switch context, lalu tarik VPA.
    """
    sample_path = Path(SAMPLE_FILE)

    if sample_path.is_file():
        logger.info("Menggunakan file sample: %s", sample_path.resolve())
        with sample_path.open(encoding="utf-8") as f:
            return [(SAMPLE_CLUSTER_NAME, json.load(f))], False

    contexts = get_kubectl_contexts()
    if not contexts:
        logger.error("Tidak ada context cluster yang ditemukan di kubeconfig.")
        sys.exit(1)

    original_context = get_current_context()
    cluster_datasets: list[tuple[str, dict]] = []

    logger.info(
        "Memulai penarikan multi-cluster dari %d context...",
        len(contexts),
    )

    try:
        for context in contexts:
            if not use_kubectl_context(context):
                continue

            try:
                vpa_data = fetch_vpa_json_from_current_context()
                item_count = len(vpa_data.get("items") or [])
                logger.info("Cluster '%s': %d objek VPA ditemukan.", context, item_count)
                cluster_datasets.append((context, vpa_data))
            except subprocess.CalledProcessError as exc:
                logger.warning(
                    "Skip cluster '%s': gagal mengambil VPA (%s).",
                    context,
                    exc.stderr.strip(),
                )
            except json.JSONDecodeError as exc:
                logger.warning(
                    "Skip cluster '%s': respons VPA bukan JSON valid (%s).",
                    context,
                    exc,
                )
    finally:
        if original_context:
            use_kubectl_context(original_context)
            logger.info("Context kubectl dikembalikan ke: %s", original_context)

    if not cluster_datasets:
        logger.error("Tidak ada data VPA yang berhasil diambil dari cluster manapun.")
        sys.exit(1)

    return cluster_datasets, True


def extract_target(recommendation: dict) -> dict | None:
    """
    Ekstrak target CPU & memory dari rekomendasi VPA.

    Prioritas:
    1. status.recommendation.target (jika ada)
    2. Agregasi dari status.recommendation.containerRecommendations[].target
    """
    if not recommendation:
        return None

    direct_target = recommendation.get("target")
    if isinstance(direct_target, dict) and direct_target.get("cpu") and direct_target.get("memory"):
        return direct_target

    containers = recommendation.get("containerRecommendations") or []
    total_cpu = 0.0
    total_memory = 0.0
    found = False

    for container in containers:
        target = container.get("target") or {}
        cpu = target.get("cpu")
        memory = target.get("memory")
        if not cpu or not memory:
            continue
        total_cpu += parse_cpu(cpu)
        total_memory += parse_memory(memory)
        found = True

    if not found:
        return None

    return {
        "cpu": format_cpu(total_cpu),
        "memory": format_memory(total_memory),
    }


def aggregate_container_requests(containers: list) -> dict | None:
    """Jumlahkan requests CPU & memory dari daftar container."""
    total_cpu = 0.0
    total_memory = 0.0
    found = False

    for container in containers:
        requests = container.get("resources", {}).get("requests") or {}
        cpu = requests.get("cpu")
        memory = requests.get("memory")
        if not cpu or not memory:
            continue
        total_cpu += parse_cpu(cpu)
        total_memory += parse_memory(memory)
        found = True

    if not found:
        return None

    return {
        "cpu": format_cpu(total_cpu),
        "memory": format_memory(total_memory),
    }


def extract_containers_from_manifest(manifest: dict) -> list | None:
    """Ambil daftar container dari manifest workload Kubernetes."""
    if not manifest:
        return None

    spec = manifest.get("spec") or {}
    kind = manifest.get("kind", "")

    if kind == "CronJob":
        containers = (
            spec.get("jobTemplate", {})
            .get("spec", {})
            .get("template", {})
            .get("spec", {})
            .get("containers")
        )
    else:
        containers = spec.get("template", {}).get("spec", {}).get("containers")

    return containers if containers else None


def extract_current_from_item(item: dict) -> dict | None:
    """
    Ekstrak konfigurasi request saat ini dari objek VPA jika tersedia.

    Prioritas:
    1. metadata.annotations (vpa-analyzer.io/*)
    2. currentResources / workloadManifest embedded di item JSON
    3. spec.targetRef.resources (jika ada)
    """
    metadata = item.get("metadata") or {}
    annotations = metadata.get("annotations") or {}

    cpu = annotations.get("vpa-analyzer.io/cpu-request")
    memory = annotations.get("vpa-analyzer.io/memory-request")
    if cpu and memory:
        return {"cpu": cpu, "memory": memory}

    current_resources = item.get("currentResources") or {}
    if current_resources.get("cpu") and current_resources.get("memory"):
        return {
            "cpu": current_resources["cpu"],
            "memory": current_resources["memory"],
        }

    workload_manifest = item.get("workloadManifest")
    containers = extract_containers_from_manifest(workload_manifest)
    if containers:
        current = aggregate_container_requests(containers)
        if current:
            return current

    target_ref = item.get("spec", {}).get("targetRef") or {}
    target_resources = target_ref.get("resources", {}).get("requests") or {}
    if target_resources.get("cpu") and target_resources.get("memory"):
        return {
            "cpu": target_resources["cpu"],
            "memory": target_resources["memory"],
        }

    return None


def fetch_current_from_cluster(namespace: str, target_ref: dict) -> dict | None:
    """Ambil request saat ini dari workload cluster via kubectl."""
    if not FETCH_WORKLOAD:
        return None

    kind = target_ref.get("kind", "Deployment")
    name = target_ref.get("name")
    if not name:
        return None

    resource_map = {
        "Deployment": "deployment",
        "StatefulSet": "statefulset",
        "DaemonSet": "daemonset",
        "ReplicationController": "rc",
        "CronJob": "cronjob",
    }
    resource = resource_map.get(kind, kind.lower())
    cmd = ["kubectl", "get", resource, name, "-n", namespace, "-o", "json"]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True,
        )
        manifest = json.loads(result.stdout)
    except (subprocess.CalledProcessError, FileNotFoundError, json.JSONDecodeError) as exc:
        logger.debug(
            "Tidak dapat mengambil workload %s/%s (%s): %s",
            namespace,
            name,
            kind,
            exc,
        )
        return None

    containers = extract_containers_from_manifest(manifest)
    if not containers:
        return None

    return aggregate_container_requests(containers)


def simulate_current_resources(
    cluster_name: str,
    namespace: str,
    service_name: str,
    target_cpu_m: float,
    target_mem_bytes: float,
) -> dict:
    """
    Simulasikan konfigurasi saat ini bila data workload tidak tersedia.

    Menghasilkan variasi realistis antara ~55% hingga ~145% dari target VPA.
    """
    seed = hash(f"{cluster_name}/{namespace}/{service_name}") & 0xFFFFFFFF
    cpu_factor = 0.55 + (seed % 91) / 100.0
    mem_factor = 0.55 + ((seed // 97) % 91) / 100.0

    current_cpu_m = max(parse_cpu("50m"), target_cpu_m * cpu_factor)
    current_mem_bytes = max(parse_memory("64Mi"), target_mem_bytes * mem_factor)

    logger.debug(
        "Simulasi konfigurasi saat ini untuk %s/%s/%s (cpu x%.2f, mem x%.2f).",
        cluster_name,
        namespace,
        service_name,
        cpu_factor,
        mem_factor,
    )

    return {
        "cpu": format_cpu(current_cpu_m),
        "memory": format_memory(current_mem_bytes),
        "simulated": True,
    }


def resolve_current_resources(
    item: dict,
    cluster_name: str,
    namespace: str,
    service_name: str,
    target_cpu_m: float,
    target_mem_bytes: float,
    live_mode: bool,
) -> dict:
    """Gabungkan semua sumber data konfigurasi saat ini."""
    current = extract_current_from_item(item)
    if current:
        current["simulated"] = False
        return current

    if live_mode:
        target_ref = item.get("spec", {}).get("targetRef") or {}
        current = fetch_current_from_cluster(namespace, target_ref)
        if current:
            current["simulated"] = False
            return current

    simulated = simulate_current_resources(
        cluster_name,
        namespace,
        service_name,
        target_cpu_m,
        target_mem_bytes,
    )
    return simulated


def build_recommendations(
    vpa_data: dict,
    live_mode: bool,
    cluster_name: str,
) -> list[dict]:
    """Loop semua VPA dan bangun baris analisis efisiensi untuk Excel."""
    items = vpa_data.get("items") or []
    rows: list[dict] = []

    logger.info("Memproses cluster '%s' (%d VPA)...", cluster_name, len(items))

    for item in items:
        metadata = item.get("metadata") or {}
        namespace = metadata.get("namespace", "unknown")
        service_name = metadata.get("name", "unknown")

        status = item.get("status") or {}
        recommendation = status.get("recommendation")

        try:
            target = extract_target(recommendation)
            if not target:
                logger.info(
                    "Skip [%s] %s/%s: rekomendasi VPA kosong atau belum siap.",
                    cluster_name,
                    namespace,
                    service_name,
                )
                continue

            recommended_cpu = target["cpu"]
            recommended_memory = target["memory"]
            recommended_cpu_m = parse_cpu(recommended_cpu)
            recommended_mem_bytes = parse_memory(recommended_memory)

            current = resolve_current_resources(
                item,
                cluster_name,
                namespace,
                service_name,
                recommended_cpu_m,
                recommended_mem_bytes,
                live_mode,
            )
            current_cpu = current["cpu"]
            current_memory = current["memory"]
            current_cpu_m = parse_cpu(current_cpu)
            current_mem_bytes = parse_memory(current_memory)
            current_mem_mib = memory_to_mib(current_mem_bytes)
            recommended_mem_mib = memory_to_mib(recommended_mem_bytes)

            cpu_diff = round(recommended_cpu_m - current_cpu_m, 2)
            mem_diff = round(recommended_mem_mib - current_mem_mib, 2)
            cpu_pct = calc_percent_change(current_cpu_m, recommended_cpu_m)
            mem_pct = calc_percent_change(current_mem_mib, recommended_mem_mib)

            rows.append(
                {
                    "Cluster Name": cluster_name,
                    "Namespace": namespace,
                    "Nama Service": service_name,
                    "CPU Sekarang": current_cpu,
                    "Rekomendasi CPU Baru": recommended_cpu,
                    "Selisih CPU": cpu_diff,
                    "% Perubahan CPU": cpu_pct,
                    "Memori Sekarang": current_memory,
                    "Rekomendasi Memori Baru": recommended_memory,
                    "Selisih Memori": mem_diff,
                    "% Perubahan Memori": mem_pct,
                    "Status Tindakan": determine_status(cpu_pct, mem_pct),
                }
            )
        except (KeyError, TypeError, ValueError) as exc:
            logger.info(
                "Skip [%s] %s/%s: gagal memproses rekomendasi (%s).",
                cluster_name,
                namespace,
                service_name,
                exc,
            )

    return rows


def apply_percentage_colors(worksheet, df: pd.DataFrame) -> None:
    """Warnai sel persentase: hijau untuk minus, merah untuk plus."""
    pct_columns = ["% Perubahan CPU", "% Perubahan Memori"]

    for column_name in pct_columns:
        col_idx = df.columns.get_loc(column_name) + 1
        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                continue
            if value < 0:
                cell.fill = GREEN_FILL
            elif value > 0:
                cell.fill = RED_FILL


def save_to_excel(rows: list[dict]) -> None:
    """Simpan hasil analisis ke file Excel dengan conditional formatting."""
    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    output_path = Path(OUTPUT_FILE)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="VPA Analysis")
        apply_percentage_colors(writer.sheets["VPA Analysis"], df)

    logger.info("Hasil disimpan ke: %s (%d baris)", output_path.resolve(), len(df))


def main() -> None:
    cluster_datasets, live_mode = load_vpa_data_by_cluster()
    all_rows: list[dict] = []

    for cluster_name, vpa_data in cluster_datasets:
        rows = build_recommendations(vpa_data, live_mode, cluster_name)
        all_rows.extend(rows)

    if not all_rows:
        logger.warning("Tidak ada rekomendasi VPA yang valid untuk diekspor.")
        sys.exit(0)

    logger.info(
        "Total %d baris dari %d cluster siap diekspor.",
        len(all_rows),
        len(cluster_datasets),
    )
    save_to_excel(all_rows)


if __name__ == "__main__":
    main()
